# finance/management/commands/import_xlsx.py
import os
import re
import random
import string
import datetime
from decimal import Decimal, InvalidOperation

import django
from django.core.management.base import BaseCommand
from django.contrib.auth.models import User
from django.db import transaction

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'sms_core.settings')
django.setup()

import openpyxl
from openpyxl import load_workbook

from finance.models import (
    Student, ClassStream, StaffProfile, Subject, Teacher, FeeStructure,
    AttendanceRecord, DisciplineReport, ExamRecord, FeeInvoice, FeeReceipt,
    StudentAttendanceRecord, TeacherAttendanceRecord, LeaveApplication,
    TimetableAllocation, HomeworkAssignment, SchoolAnnouncement, SchoolAsset,
    AssetMaintenanceLog, LessonPlan, LearningMaterial, TimetableSlot,
    SchoolHoliday,
)
from finance.school_config import CBC_LEVELS


DEFAULT_EXCEL_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))),
    "PP1 Student list (6).xlsx",
)


def split_name(full_name):
    parts = full_name.strip().split()
    if not parts:
        return "", ""
    first = parts[0]
    last = " ".join(parts[1:]) if len(parts) > 1 else ""
    return first, last


def generate_password(length=10):
    chars = string.ascii_letters + string.digits
    return ''.join(random.choice(chars) for _ in range(length))


def money(value):
    """Return a workbook currency value as a decimal, treating Cleared as zero."""
    if value is None:
        return Decimal("0.00")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    cleaned = re.sub(r"[^0-9.-]", "", str(value))
    if not cleaned:
        return Decimal("0.00")
    try:
        return Decimal(cleaned).quantize(Decimal("0.01"))
    except InvalidOperation:
        return Decimal("0.00")


class Command(BaseCommand):
    help = "Imports students and staff from the Kabiero Academy Excel workbook into the Django database."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            dest="excel_path",
            default=DEFAULT_EXCEL_PATH,
            help="Path to the Excel workbook to import.",
        )
        parser.add_argument(
            "--replace",
            action="store_true",
            help="Remove all existing school operational records before importing this workbook.",
        )
        parser.add_argument(
            "--uniform-admission-numbers",
            action="store_true",
            help="Assign every imported learner a sequential KBA-2026-#### admission number.",
        )
        parser.add_argument(
            "--generate-provisional-admissions",
            action="store_true",
            help=(
                "Import rows with no admission number using PROV-<GRADE>-<ROW> IDs. "
                "Use only until official admission numbers are collected."
            ),
        )

    def handle(self, *args, **options):
        excel_path = os.path.abspath(options["excel_path"])
        if not os.path.exists(excel_path):
            self.stdout.write(self.style.ERROR(f"Excel file not found at: {excel_path}"))
            return

        wb = load_workbook(excel_path, data_only=True)

        with transaction.atomic():
            if options["replace"]:
                self.clear_school_data()
            self.import_students(
                wb,
                options["generate_provisional_admissions"],
                options["uniform_admission_numbers"],
            )
            self.import_staff(wb)
            self.import_fee_structure(wb)

        self.stdout.write(self.style.SUCCESS("Import completed successfully."))

    def clear_school_data(self):
        """Clear school records but retain superuser accounts for administrator access."""
        # Delete dependent records first; cascades then cover the remaining links.
        for model in (
            AssetMaintenanceLog, AttendanceRecord, DisciplineReport, ExamRecord,
            FeeReceipt, FeeInvoice, StudentAttendanceRecord, TeacherAttendanceRecord,
            LeaveApplication, TimetableAllocation, HomeworkAssignment, LessonPlan,
            LearningMaterial, TimetableSlot, SchoolAnnouncement, SchoolHoliday,
            SchoolAsset, Student, Teacher, StaffProfile, Subject, FeeStructure,
            ClassStream,
        ):
            model.objects.all().delete()
        User.objects.filter(is_superuser=False).delete()
        self.stdout.write(self.style.WARNING("Existing school operational data removed; superuser accounts retained."))

    def import_students(
        self, wb, generate_provisional_admissions=False, uniform_admission_numbers=False
    ):
        if "Data Entry" not in wb.sheetnames:
            self.stdout.write(self.style.WARNING("Sheet 'Data Entry' not found."))
            return

        ws = wb["Data Entry"]
        student_count = 0
        stream_count = 0
        admission_sequence = 0

        for row_number, row in enumerate(ws.iter_rows(min_row=6, values_only=True), start=6):
            adm_no = row[1]
            name = row[2]
            grade = row[3]
            balance_raw = row[4]
            parent_name = row[5]
            parent_phone = row[6]

            if name is None or grade is None:
                continue

            if isinstance(adm_no, float):
                adm_no = str(int(adm_no))
            else:
                adm_no = str(adm_no).strip() if adm_no is not None else ""
            name = str(name).strip()
            grade = str(grade).strip()
            balance_raw = balance_raw if balance_raw is not None else 0
            parent_name = str(parent_name).strip() if parent_name is not None else ""
            parent_phone = str(parent_phone).strip() if parent_phone is not None else ""

            if not name or not grade:
                continue

            grade = grade.strip()
            if grade not in CBC_LEVELS:
                self.stdout.write(self.style.WARNING(f"Skipping unknown grade '{grade}' for adm {adm_no}"))
                continue

            admission_sequence += 1
            if uniform_admission_numbers:
                adm_no = f"KBA-2026-{admission_sequence:04d}"
            elif not adm_no:
                if not generate_provisional_admissions:
                    self.stdout.write(self.style.WARNING(
                        f"Skipping {name}: admission number is missing. "
                        "Re-run with --generate-provisional-admissions to import it provisionally."
                    ))
                    continue
                grade_token = re.sub(r"[^A-Z0-9]", "", grade.upper())
                adm_no = f"PROV-{grade_token}-{row_number:03d}"

            stream_instance, created = ClassStream.objects.get_or_create(name=grade)
            if created:
                stream_count += 1

            first_name, last_name = split_name(name)

            balance = money(balance_raw)

            student, created = Student.objects.update_or_create(
                admission_number=adm_no,
                defaults={
                    "first_name": first_name,
                    "last_name": last_name,
                    "class_stream": stream_instance,
                    "guardian_name": parent_name or "Not Provided",
                    "parent_phone": parent_phone or "0700000000",
                    "current_balance": balance,
                    "status": "ACTIVE",
                    "is_active": True,
                },
            )
            if created:
                student_count += 1

        self.stdout.write(self.style.SUCCESS(f"Students: {student_count} created, streams: {stream_count} created."))

    def import_staff(self, wb):
        if "Staff Directory" not in wb.sheetnames:
            self.stdout.write(self.style.WARNING("Sheet 'Staff Directory' not found."))
            return

        ws = wb["Staff Directory"]
        staff_count = 0

        for row_number, row in enumerate(ws.iter_rows(min_row=6, values_only=True), start=6):
            no_col = row[1]
            staff_id = row[2]
            full_name = row[3]
            gender = row[4]
            phone = row[5]
            position = row[6]
            subject_dept = row[7]
            employment_type = row[8]
            date_joined = row[9]
            status_raw = row[10]

            if not full_name:
                continue

            full_name = str(full_name).strip()
            if not full_name:
                continue

            if isinstance(staff_id, float):
                staff_id = str(int(staff_id))
            else:
                staff_id = str(staff_id).strip() if staff_id is not None else ""
            if staff_id in ["—", "-", "None", ""]:
                staff_id = ""

            phone = str(phone).strip() if phone is not None else ""
            if phone in ["—", "-", "None", ""]:
                phone = ""

            position = str(position).strip() if position is not None else ""
            subject_dept = str(subject_dept).strip() if subject_dept is not None else ""
            status_raw = str(status_raw).strip() if status_raw is not None else ""

            first_name, last_name = split_name(full_name)
            if not last_name:
                last_name = "Staff"

            employee_number = staff_id if staff_id else f"EMP-{(no_col or row_number - 5):03d}"
            phone_line = phone if phone else "0700000000"

            role_map = {
                "Administrator": "PRINCIPAL",
                "Teacher": "TEACHER",
                "Support Staff": "SUPPORT",
            }
            role_designation = role_map.get(position, "SUPPORT")

            current_status = "ACTIVE"
            if status_raw:
                if status_raw.lower() in ["on leave", "suspended"]:
                    current_status = status_raw.upper().replace(" ", "_")

            username_base = re.sub(r"[^a-z0-9]", "", full_name.lower()[:20])
            username = username_base if username_base else f"staff{random.randint(1000, 9999)}"
            if User.objects.filter(username=username).exists():
                username = f"{username}{no_col or row_number - 5}"
            password = generate_password()

            user, user_created = User.objects.get_or_create(
                username=username,
                defaults={
                    "first_name": first_name,
                    "last_name": last_name,
                    "email": "",
                },
            )
            if user_created:
                user.set_password(password)
                user.save()
                self.stdout.write(self.style.NOTICE(f"Created user {username} with password {password}"))

            specialization = subject_dept if subject_dept and subject_dept != "—" else "General"

            profile, created = StaffProfile.objects.update_or_create(
                user=user,
                defaults={
                    "employee_number": employee_number,
                    "role_designation": role_designation,
                    "phone_line": phone_line,
                    "specialization": specialization,
                    "base_salary_kes": money(row[11]),
                    "current_status": current_status,
                    "performance_score": 85,
                },
            )
            if created:
                staff_count += 1

        self.stdout.write(self.style.SUCCESS(f"Staff: {staff_count} profiles created."))

    def import_fee_structure(self, wb):
        """Import the 2026 mandatory term totals represented by the supplied schedule.

        The current model stores one amount per grade and term, so optional lunch is
        intentionally excluded.  School fees, parental award, medical and activity
        charges are combined where applicable.
        """
        if "Fee Structure" not in wb.sheetnames:
            self.stdout.write(self.style.WARNING("Sheet 'Fee Structure' not found."))
            return

        mandatory = {
            "Playgroup": (400, 400, 400),
            "PP1": (4200, 3200, 2200),
            "PP2": (4200, 3200, 2200),
            "Grade 1": (5400, 4400, 2200),
            "Grade 2": (5400, 4400, 2200),
            "Grade 3": (5400, 4400, 2200),
            "Grade 4": (5400, 4400, 2200),
            "Grade 5": (5400, 4400, 2200),
            "Grade 6": (5400, 4400, 2200),
            "Grade 7": (6100, 5100, 2900),
            "Grade 8": (6100, 5100, 2900),
            "Grade 9": (6100, 5100, 2900),
        }
        for level, amounts in mandatory.items():
            for term, amount in zip(("TERM_1", "TERM_2", "TERM_3"), amounts):
                FeeStructure.objects.update_or_create(
                    level=level, term=term, year=2026,
                    defaults={"amount": Decimal(str(amount))},
                )
        self.stdout.write(self.style.SUCCESS("Fee structure: 36 mandatory 2026 term rates imported (optional lunch excluded)."))
